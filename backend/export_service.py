"""
Servizio per esportazione dati di produzione e KPI
Supporta formati: CSV, Excel, JSON
"""

import csv
import json
from collections import Counter
from io import StringIO, BytesIO
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference

from database import DatabaseRepository


class ExportService:
    """
    Servizio per generare export dati produzione in vari formati
    """
    
    def __init__(self, db: DatabaseRepository):
        self.db = db
    
    # ========================================================================
    # RACCOLTA DATI
    # ========================================================================
    
    async def get_dati_produzione(
        self, 
        data_inizio: str, 
        data_fine: str
    ) -> Dict[str, Any]:
        """
        Raccoglie tutti i dati di produzione per il periodo specificato
        
        Args:
            data_inizio: Data inizio periodo (formato ISO: YYYY-MM-DD)
            data_fine: Data fine periodo (formato ISO: YYYY-MM-DD)
            
        Returns:
            Dizionario con tutti i dati aggregati
        """
        async with self.db.db.execute(
            """SELECT 
                c.id,
                c.data_ordine,
                c.data_inizio_produzione,
                c.data_fine_produzione,
                c.quantita_richiesta,
                c.quantita_prodotta,
                c.stato,
                cl.nome as cliente_nome,
                r.nome as ricetta_nome,
                c.priorita
            FROM commesse c
            LEFT JOIN clienti cl ON c.cliente_id = cl.id
            LEFT JOIN ricette r ON c.ricetta_id = r.id
            WHERE date(c.data_ordine) BETWEEN date(?) AND date(?)
            ORDER BY c.data_ordine DESC""",
            (data_inizio, data_fine)
        ) as cursor:
            commesse_rows = await cursor.fetchall()
            
        # Converti in lista di dizionari
        commesse = []
        for row in commesse_rows:
            commesse.append({
                'id': row[0],
                'data_ordine': row[1],
                'data_inizio': row[2],
                'data_fine': row[3],
                'quantita_richiesta': row[4],
                'quantita_prodotta': row[5],
                'stato': row[6],
                'cliente': row[7],
                'ricetta': row[8],
                'priorita': row[9],
                'percentuale_completamento': round((row[5] / row[4] * 100) if row[4] > 0 else 0, 2)
            })
        
        # Calcola durata produzione per commesse completate
        for comm in commesse:
            if comm['data_inizio'] and comm['data_fine']:
                inizio = datetime.fromisoformat(comm['data_inizio'])
                fine = datetime.fromisoformat(comm['data_fine'])
                durata = fine - inizio
                comm['durata_ore'] = round(durata.total_seconds() / 3600, 2)
                comm['pezzi_ora'] = round(comm['quantita_prodotta'] / comm['durata_ore'], 2) if comm['durata_ore'] > 0 else 0
            else:
                comm['durata_ore'] = None
                comm['pezzi_ora'] = None
        
        # Statistiche allarmi nel periodo
        async with self.db.db.execute(
            """SELECT 
                codice_allarme,
                COUNT(*) as occorrenze,
                AVG(durata_secondi) as durata_media_sec,
                SUM(durata_secondi) as durata_totale_sec
            FROM allarmi_storico
            WHERE date(timestamp_inizio) BETWEEN date(?) AND date(?)
            AND durata_secondi IS NOT NULL
            GROUP BY codice_allarme
            ORDER BY occorrenze DESC""",
            (data_inizio, data_fine)
        ) as cursor:
            allarmi_rows = await cursor.fetchall()
        
        allarmi = []
        for row in allarmi_rows:
            allarmi.append({
                'codice': row[0],
                'occorrenze': row[1],
                'durata_media_minuti': round(row[2] / 60, 2) if row[2] else 0,
                'durata_totale_ore': round(row[3] / 3600, 2) if row[3] else 0
            })
        
        # Eventi macchina nel periodo
        async with self.db.db.execute(
            """SELECT 
                tipo_evento,
                COUNT(*) as conteggio
            FROM eventi_macchina
            WHERE date(timestamp) BETWEEN date(?) AND date(?)
            GROUP BY tipo_evento
            ORDER BY conteggio DESC""",
            (data_inizio, data_fine)
        ) as cursor:
            eventi_rows = await cursor.fetchall()
        
        eventi = [{'tipo': row[0], 'conteggio': row[1]} for row in eventi_rows]

        # Sessioni pannello nel periodo (senza commessa, per export)
        async with self.db.db.execute(
            """SELECT id, timestamp_inizio, timestamp_fine, durata_secondi,
                      ricetta_nome, quantita_prodotta, contatore_lotto
               FROM sessioni_produzione
               WHERE commessa_id IS NULL AND stato = 'chiusa'
                 AND date(timestamp_inizio) BETWEEN date(?) AND date(?)
               ORDER BY timestamp_inizio DESC""",
            (data_inizio, data_fine)
        ) as cursor:
            sessioni_rows = await cursor.fetchall()

        sessioni_pannello = []
        for row in sessioni_rows:
            durata_ore = round(row[3] / 3600, 2) if row[3] else None
            sessioni_pannello.append({
                'id': row[0], 'timestamp_inizio': row[1], 'timestamp_fine': row[2],
                'durata_ore': durata_ore, 'ricetta_nome': row[4],
                'quantita_prodotta': row[5], 'contatore_lotto': row[6],
                'pezzi_ora': round(row[5] / durata_ore, 2) if durata_ore and durata_ore > 0 else 0,
            })

        return {
            'periodo': {
                'inizio': data_inizio,
                'fine': data_fine
            },
            'commesse': commesse,
            'allarmi': allarmi,
            'eventi': eventi,
            'sessioni_pannello': sessioni_pannello,
            'timestamp_export': datetime.now().isoformat()
        }
    
    async def calcola_tempo_effettivo_macchina(
        self,
        data_inizio: str,
        data_fine: str
    ) -> Dict[str, Any]:
        """
        Calcola il tempo effettivo di utilizzo della macchina basato su eventi reali
        
        Logica:
        - Cerca eventi START (macchina accesa)
        - Calcola durata fino a STOP o cambio stato
        - Ignora periodi senza attività
        
        Returns:
            Dizionario con tempi effettivi calcolati
        """
        # Recupera tutti gli eventi macchina ordinati cronologicamente
        async with self.db.db.execute(
            """SELECT 
                timestamp,
                tipo_evento,
                stato_macchina,
                lavorazione_id
            FROM eventi_macchina
            WHERE date(timestamp) BETWEEN date(?) AND date(?)
            ORDER BY timestamp ASC""",
            (data_inizio, data_fine)
        ) as cursor:
            eventi = await cursor.fetchall()
        
        # Stati che indicano macchina operativa
        stati_operativi = ['START_AUTOMATICO', 'START_MANUALE']
        stati_fermo = ['STOP_AUTOMATICO', 'STOP_MANUALE', 'EMERGENZA']
        
        tempo_operativo_secondi = 0
        tempo_fermo_operativo_secondi = 0  # Fermo DURANTE utilizzo (es. allarmi)
        
        ultimo_stato = None
        ultimo_timestamp = None
        
        for evento in eventi:
            timestamp_str = evento[0]
            tipo_evento = evento[1]
            stato = evento[2]
            
            timestamp = datetime.fromisoformat(timestamp_str)
            
            # Se avevamo uno stato precedente, calcola la durata
            if ultimo_stato and ultimo_timestamp:
                durata_secondi = (timestamp - ultimo_timestamp).total_seconds()
                
                if ultimo_stato in stati_operativi:
                    tempo_operativo_secondi += durata_secondi
                elif ultimo_stato in stati_fermo and tipo_evento in ['ALLARME_INIZIO', 'ALLARME_FINE']:
                    # Fermo per allarme durante periodo operativo
                    tempo_fermo_operativo_secondi += durata_secondi
            
            # Aggiorna stato
            if stato:
                ultimo_stato = stato
                ultimo_timestamp = timestamp
        
        # Converti in ore
        ore_operative = round(tempo_operativo_secondi / 3600, 2)
        ore_fermo_operative = round(tempo_fermo_operativo_secondi / 3600, 2)
        
        return {
            'ore_macchina_accesa': ore_operative,
            'ore_fermo_durante_utilizzo': ore_fermo_operative,
            'ore_nette_produzione': round(ore_operative - ore_fermo_operative, 2)
        }

    async def calcola_kpi(
        self,
        data_inizio: str,
        data_fine: str
    ) -> Dict[str, Any]:
        """
        Calcola i KPI principali per il periodo specificato

        Returns:
            Dizionario con KPI calcolati
        """
        dati = await self.get_dati_produzione(data_inizio, data_fine)
        
        commesse = dati['commesse']
        stato_counts = Counter(c['stato'] for c in commesse)
        commesse_completate = [c for c in commesse if c['stato'] == 'completata']
        
        # KPI Produzione
        totale_pezzi_prodotti = sum(c['quantita_prodotta'] for c in commesse)
        totale_pezzi_richiesti = sum(c['quantita_richiesta'] for c in commesse)
        
        # Calcola tempo totale di produzione (solo commesse completate)
        tempo_produzione_ore = sum(
            c['durata_ore'] for c in commesse_completate 
            if c['durata_ore'] is not None
        )
        
        # Performance
        pezzi_ora_medio = round(
            totale_pezzi_prodotti / tempo_produzione_ore, 2
        ) if tempo_produzione_ore > 0 else 0
        
        # Efficienza completamento
        tasso_completamento = round(
            len(commesse_completate) / len(commesse) * 100, 2
        ) if len(commesse) > 0 else 0
        
        # Tempo medio per commessa
        tempo_medio_commessa = round(
            tempo_produzione_ore / len(commesse_completate), 2
        ) if len(commesse_completate) > 0 else 0
        
        # Allarmi
        totale_allarmi = sum(a['occorrenze'] for a in dati['allarmi'])
        tempo_fermo_allarmi_ore = sum(a['durata_totale_ore'] for a in dati['allarmi'])

        # Allarmi durante commesse vs fuori commesse
        async with self.db.db.execute(
            """SELECT
                COALESCE(SUM(CASE WHEN lavorazione_id IS NOT NULL THEN durata_secondi ELSE 0 END), 0) as durante_commesse,
                COALESCE(SUM(CASE WHEN lavorazione_id IS NULL THEN durata_secondi ELSE 0 END), 0) as fuori_commesse
            FROM allarmi_storico
            WHERE date(timestamp_inizio) BETWEEN date(?) AND date(?)
            AND durata_secondi IS NOT NULL""",
            (data_inizio, data_fine)
        ) as cursor:
            row = await cursor.fetchone()
            ore_allarmi_durante_commesse = round(row[0] / 3600, 2) if row[0] else 0
            ore_allarmi_fuori_commesse = round(row[1] / 3600, 2) if row[1] else 0

        # Sessioni pannello senza commessa (nessun doppio conteggio)
        async with self.db.db.execute(
            """SELECT
                COUNT(*),
                COALESCE(SUM(quantita_prodotta), 0),
                COALESCE(SUM(durata_secondi), 0)
            FROM sessioni_produzione
            WHERE commessa_id IS NULL
              AND stato = 'chiusa'
              AND date(timestamp_inizio) BETWEEN date(?) AND date(?)""",
            (data_inizio, data_fine)
        ) as cursor:
            row_sess = await cursor.fetchone()
            sessioni_pannello_count = row_sess[0]
            pezzi_pannello = row_sess[1]
            ore_pannello = round(row_sess[2] / 3600, 2)

        # Calcola giorni calendario
        giorni_periodo = (datetime.fromisoformat(data_fine) - datetime.fromisoformat(data_inizio)).days + 1
        
        # Calcola giorni effettivi lavorati (con almeno 1 evento)
        async with self.db.db.execute(
            """SELECT COUNT(DISTINCT date(timestamp)) 
            FROM eventi_macchina
            WHERE date(timestamp) BETWEEN date(?) AND date(?)""",
            (data_inizio, data_fine)
        ) as cursor:
            giorni_lavorati = (await cursor.fetchone())[0]
        
        return {
            'periodo': {
                'inizio': data_inizio,
                'fine': data_fine,
                'giorni_calendario': giorni_periodo,
                'giorni_lavorati': giorni_lavorati,
                'giorni_inattivi': giorni_periodo - giorni_lavorati
            },
            'produzione': {
                'pezzi_prodotti': totale_pezzi_prodotti,
                'pezzi_richiesti': totale_pezzi_richiesti,
                'tasso_completamento_pezzi_perc': round(totale_pezzi_prodotti / totale_pezzi_richiesti * 100, 2) if totale_pezzi_richiesti > 0 else 0,
                'pezzi_ora_medio': pezzi_ora_medio,
                'tempo_produzione_totale_ore': round(tempo_produzione_ore, 2)
            },
            'commesse': {
                'totali': len(commesse),
                'completate': len(commesse_completate),
                'in_corso': stato_counts['in_lavorazione'],
                'in_attesa': stato_counts['in_attesa'],
                'tasso_completamento_perc': tasso_completamento,
                'tempo_medio_commessa_ore': tempo_medio_commessa
            },
            'tempi': {
                'ore_produzione_commesse': round(tempo_produzione_ore, 2),
                'ore_fermo_durante_commesse': ore_allarmi_durante_commesse,
                'ore_fermo_fuori_commesse': ore_allarmi_fuori_commesse,
                'ore_nette': round(tempo_produzione_ore - ore_allarmi_durante_commesse, 2)
            },
            'allarmi': {
                'totale_occorrenze': totale_allarmi,
                'tempo_fermo_totale_ore': round(tempo_fermo_allarmi_ore, 2),
                'allarmi_per_giorno_lavorato': round(totale_allarmi / giorni_lavorati, 2) if giorni_lavorati > 0 else 0
            },
            'sessioni_pannello': {
                'totale': sessioni_pannello_count,
                'pezzi_prodotti': pezzi_pannello,
                'ore_produzione': ore_pannello,
                'pezzi_ora_medio': round(pezzi_pannello / ore_pannello, 2) if ore_pannello > 0 else 0,
            },
            'totale': {
                'pezzi_prodotti': totale_pezzi_prodotti + pezzi_pannello,
                'ore_produzione': round(tempo_produzione_ore + ore_pannello, 2),
            }
        }
    
    # ========================================================================
    # EXPORT CSV
    # ========================================================================
    
    async def export_csv(
        self, 
        data_inizio: str, 
        data_fine: str
    ) -> str:
        """
        Genera CSV con dati produzione
        
        Returns:
            String contenente il CSV
        """
        dati = await self.get_dati_produzione(data_inizio, data_fine)
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            'ID Commessa', 'Cliente', 'Ricetta', 'Data Ordine', 
            'Data Inizio', 'Data Fine', 'Stato', 'Priorità',
            'Quantità Richiesta', 'Quantità Prodotta', 'Completamento %',
            'Durata (ore)', 'Pezzi/Ora'
        ])
        
        # Dati
        for c in dati['commesse']:
            writer.writerow([
                c['id'],
                c['cliente'],
                c['ricetta'],
                c['data_ordine'],
                c['data_inizio'] or '-',
                c['data_fine'] or '-',
                c['stato'],
                c['priorita'],
                c['quantita_richiesta'],
                c['quantita_prodotta'],
                c['percentuale_completamento'],
                c['durata_ore'] or '-',
                c['pezzi_ora'] or '-'
            ])
        
        # Sezione allarmi
        writer.writerow([])
        writer.writerow(['ALLARMI NEL PERIODO'])
        writer.writerow(['Codice', 'Occorrenze', 'Durata Media (min)', 'Durata Totale (ore)'])

        for a in dati['allarmi']:
            writer.writerow([
                a['codice'],
                a['occorrenze'],
                a['durata_media_minuti'],
                a['durata_totale_ore']
            ])

        # Sezione sessioni pannello
        if dati['sessioni_pannello']:
            writer.writerow([])
            writer.writerow(['SESSIONI PANNELLO (senza commessa)'])
            writer.writerow(['ID', 'Ricetta', 'Inizio', 'Fine', 'Durata (ore)', 'Pezzi', 'Pezzi/Ora'])
            for s in dati['sessioni_pannello']:
                writer.writerow([
                    s['id'], s['ricetta_nome'], s['timestamp_inizio'],
                    s['timestamp_fine'] or '-', s['durata_ore'] or '-',
                    s['quantita_prodotta'], s['pezzi_ora']
                ])

        return output.getvalue()
    
    # ========================================================================
    # EXPORT EXCEL
    # ========================================================================
    
    async def export_excel(
        self, 
        data_inizio: str, 
        data_fine: str
    ) -> BytesIO:
        """
        Genera Excel con dati produzione e grafici
        
        Returns:
            BytesIO contenente il file Excel
        """
        dati = await self.get_dati_produzione(data_inizio, data_fine)
        kpi = await self.calcola_kpi(data_inizio, data_fine)
        
        # Crea workbook
        wb = openpyxl.Workbook()
        
        # === FOGLIO 1: KPI ===
        ws_kpi = wb.active
        ws_kpi.title = "KPI"
        
        # Stile header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Titolo
        ws_kpi['A1'] = f"REPORT KPI PRODUZIONE"
        ws_kpi['A1'].font = Font(bold=True, size=14)
        ws_kpi['A2'] = f"Periodo: {data_inizio} - {data_fine}"
        
        row = 4
        
        # KPI Produzione
        ws_kpi[f'A{row}'] = "PRODUZIONE"
        ws_kpi[f'A{row}'].font = header_font
        ws_kpi[f'A{row}'].fill = header_fill
        row += 1
        
        for key, value in kpi['produzione'].items():
            ws_kpi[f'A{row}'] = key.replace('_', ' ').title()
            ws_kpi[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # KPI Commesse
        ws_kpi[f'A{row}'] = "COMMESSE"
        ws_kpi[f'A{row}'].font = header_font
        ws_kpi[f'A{row}'].fill = header_fill
        row += 1
        
        for key, value in kpi['commesse'].items():
            ws_kpi[f'A{row}'] = key.replace('_', ' ').title()
            ws_kpi[f'B{row}'] = value
            row += 1
        
        row += 1

        # KPI Tempi
        ws_kpi[f'A{row}'] = "TEMPI PRODUZIONE"
        ws_kpi[f'A{row}'].font = header_font
        ws_kpi[f'A{row}'].fill = header_fill
        row += 1

        for key, value in kpi['tempi'].items():
            ws_kpi[f'A{row}'] = key.replace('_', ' ').title()
            ws_kpi[f'B{row}'] = value
            row += 1

        row += 1

        # KPI Sessioni Pannello
        ws_kpi[f'A{row}'] = "SESSIONI PANNELLO"
        ws_kpi[f'A{row}'].font = header_font
        ws_kpi[f'A{row}'].fill = header_fill
        row += 1

        for key, value in kpi['sessioni_pannello'].items():
            ws_kpi[f'A{row}'] = key.replace('_', ' ').title()
            ws_kpi[f'B{row}'] = value
            row += 1

        row += 1

        # KPI Totale
        ws_kpi[f'A{row}'] = "TOTALE (commesse + pannello)"
        ws_kpi[f'A{row}'].font = header_font
        ws_kpi[f'A{row}'].fill = header_fill
        row += 1

        for key, value in kpi['totale'].items():
            ws_kpi[f'A{row}'] = key.replace('_', ' ').title()
            ws_kpi[f'B{row}'] = value
            row += 1

        row += 1

        # Formatta colonne
        ws_kpi.column_dimensions['A'].width = 35
        ws_kpi.column_dimensions['B'].width = 20
        
        # === FOGLIO 2: Commesse ===
        ws_comm = wb.create_sheet("Commesse")
        
        # Header
        headers = [
            'ID', 'Cliente', 'Ricetta', 'Data Ordine', 'Data Inizio', 'Data Fine',
            'Stato', 'Priorità', 'Q.tà Richiesta', 'Q.tà Prodotta', 
            'Completamento %', 'Durata (ore)', 'Pezzi/Ora'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_comm.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Dati
        for row_idx, comm in enumerate(dati['commesse'], 2):
            ws_comm.cell(row_idx, 1, comm['id'])
            ws_comm.cell(row_idx, 2, comm['cliente'])
            ws_comm.cell(row_idx, 3, comm['ricetta'])
            ws_comm.cell(row_idx, 4, comm['data_ordine'])
            ws_comm.cell(row_idx, 5, comm['data_inizio'] or '-')
            ws_comm.cell(row_idx, 6, comm['data_fine'] or '-')
            ws_comm.cell(row_idx, 7, comm['stato'])
            ws_comm.cell(row_idx, 8, comm['priorita'])
            ws_comm.cell(row_idx, 9, comm['quantita_richiesta'])
            ws_comm.cell(row_idx, 10, comm['quantita_prodotta'])
            ws_comm.cell(row_idx, 11, comm['percentuale_completamento'])
            ws_comm.cell(row_idx, 12, comm['durata_ore'] or '-')
            ws_comm.cell(row_idx, 13, comm['pezzi_ora'] or '-')
        
        # Auto-width
        for col in ws_comm.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_comm.column_dimensions[column].width = min(max_length + 2, 30)
        
        # === FOGLIO 3: Allarmi ===
        ws_alarm = wb.create_sheet("Allarmi")
        
        headers_alarm = ['Codice', 'Occorrenze', 'Durata Media (min)', 'Durata Totale (ore)']
        for col, header in enumerate(headers_alarm, 1):
            cell = ws_alarm.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, alarm in enumerate(dati['allarmi'], 2):
            ws_alarm.cell(row_idx, 1, alarm['codice'])
            ws_alarm.cell(row_idx, 2, alarm['occorrenze'])
            ws_alarm.cell(row_idx, 3, alarm['durata_media_minuti'])
            ws_alarm.cell(row_idx, 4, alarm['durata_totale_ore'])
        
        for col in ws_alarm.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_alarm.column_dimensions[column].width = max_length + 2

        # === FOGLIO 4: Sessioni Pannello ===
        ws_sess = wb.create_sheet("Sessioni Pannello")

        headers_sess = ['ID', 'Ricetta', 'Inizio', 'Fine', 'Durata (ore)', 'Pezzi Prodotti', 'Pezzi/Ora']
        for col, header in enumerate(headers_sess, 1):
            cell = ws_sess.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, s in enumerate(dati['sessioni_pannello'], 2):
            ws_sess.cell(row_idx, 1, s['id'])
            ws_sess.cell(row_idx, 2, s['ricetta_nome'])
            ws_sess.cell(row_idx, 3, s['timestamp_inizio'])
            ws_sess.cell(row_idx, 4, s['timestamp_fine'] or '-')
            ws_sess.cell(row_idx, 5, s['durata_ore'] or '-')
            ws_sess.cell(row_idx, 6, s['quantita_prodotta'])
            ws_sess.cell(row_idx, 7, s['pezzi_ora'])

        for col in ws_sess.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws_sess.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

        # Salva in BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    # ========================================================================
    # EXPORT JSON
    # ========================================================================
    
    async def export_json(
        self, 
        data_inizio: str, 
        data_fine: str,
        include_kpi: bool = True
    ) -> str:
        """
        Genera JSON con dati produzione
        
        Args:
            include_kpi: Se True, include anche i KPI calcolati
            
        Returns:
            String JSON
        """
        dati = await self.get_dati_produzione(data_inizio, data_fine)
        
        if include_kpi:
            kpi = await self.calcola_kpi(data_inizio, data_fine)
            dati['kpi'] = kpi
        
        return json.dumps(dati, indent=2, ensure_ascii=False)